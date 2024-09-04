from openpyxl import load_workbook
import re
import pandas as pd
import psycopg2
import uuid
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time


conn = psycopg2.connect(
    dbname="postgres",
    user="postgres",
    password="12345678",
    host="localhost",
    port="5432"
)
cur = conn.cursor()
doc_type='10K'

def download_excel(ticker,quarter,year,doc_url):
    print(doc_url)
    download_dir = r"C:\Users\dell\OneDrive\Documents\10-k 10-q extraction"


    # Configure Chrome options
    options = webdriver.ChromeOptions()
    options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,  # Disable download prompt
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True  # Enable safe browsing
    })

    # chrome_driver_path = "C:\\Users\\Parv Malhotra\\Documents\\drivers\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe"
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    x = doc_url.split('/')
    cik = x[6]
    ac_no = x[7]
    ac_no = ac_no[0:10] + "-" + ac_no[10:12] + "-" + ac_no[12:]

    print(ac_no)
    cgi_url = f'https://www.sec.gov/cgi-bin/viewer?action=view&cik={cik}&accession_number={ac_no}&xbrl_type=v#'
    try:
        driver.get(cgi_url)
        print(driver.title)

        buttons = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'xbrlviewer'))
        )
        print(buttons[1].text)
        file_path = f"Financial_Report_{doc_type}{ticker}{quarter}_{year}.xlsx"
        driver.execute_script(f"arguments[0].setAttribute('download', '{file_path}')", buttons[1])
        buttons[1].click()

        time.sleep(10)  # Wait for the download to complete

    finally:
        driver.quit()

def correct_values(cashflowTitle,st, parent_index, is_parent, unit):
    # try:
    sheet_title=""
    print(cashflowTitle)
    if st==1:
        sheet_title="Cash Flow Statement"
    elif st==2:
        sheet_title="Balance Sheet"
    else:
        sheet_title="Income Statement"
    sheet = workbook[value_to_sheet_title[cashflowTitle]]
    parent = ""
    parent_count = 0
    child_count = 0
    unique_id = None
    unique_id_independent = uuid.uuid4()
    flag=0
    content=[]
    data={}
    val_index=1
    
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == 1:
            if row[1]==None:
                val_index=2
            continue

        if row[0] is None or (isinstance(row[0], str) and row[0].strip().lower() == 'none'):
            continue
        # print("row [0]",row[0])
        if "Abstract".lower() in row[0]:
            continue
        if "Commitments and contingencies".lower() in row[0].lower():
                continue
        if st==2 and "total" in row[0].lower():
            flag=1
        if len(row[0])>255:
            continue
        # print("INDEX")
        if(parent_index[index] >1):
            continue
        
        normalized_row = [cell.replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]
        content.append(normalized_row)
        if all(cell is None or cell == '' for cell in normalized_row[1:]):
            
            flag=0
            if child_count>0:
                parent=""
            
            parent+=row[0]
                
            # parent = row[0]
            
            #LENGTH EXCEEDING CASE
            if len(parent)>255:
                break
            
            unique_id = uuid.uuid4()
            if index in is_parent:
                prefix_parent = sheet.cell( row = is_parent[index], column=1).value
                if prefix_parent != None:
                    parent = prefix_parent+ "-" + parent
            print("parent:", parent)
            
            cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s, %s);
            """, (ticker, quarter, year, sheet_title,doc_type, parent, str(unique_id), parent_count, unit))
            
            parent_count += 1
            child_count = 0
        else:
            child = row[0]
            if parent != "":
                # if flag==1 and "long term" not in child.lower() and "total" not in child.lower():
                #         child="Long Term "+child
                
                cur.execute("""
                    INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (heading_id, sub_heading) 
                    WHERE sub_heading IS NOT NULL 
                    DO UPDATE SET
                        field_data = EXCLUDED.field_data,
                        ordering_index = EXCLUDED.ordering_index;
                """, (str(unique_id), child, row[val_index], child_count))
                
                if st==2 and "total" in row[0].lower() and "current" in parent.lower():
                    parent_count+=1
                    parent="Long Term "+parent
                    unique_id = uuid.uuid4()
                    cur.execute("""
                        INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                        VALUES (%s, %s, %s, %s, %s, %s,%s, %s,%s);
                    """, (ticker, quarter, year, sheet_title,doc_type, parent, str(unique_id), parent_count, unit))
                    
                if st==0 and "total" in row[0].lower() and "operating" in parent.lower():
                    parent_count+=1
                    parent="Non "+parent
                    unique_id = uuid.uuid4()
                    cur.execute("""
                        INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                        VALUES (%s, %s, %s, %s, %s, %s,%s, %s,%s);
                    """, (ticker, quarter, year, sheet_title,doc_type, parent, str(unique_id), parent_count, unit))
                    
                child_count += 1

            else:
                
                unique_id = uuid.uuid4()
                
                cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s,%s);
            """, (ticker, quarter, year, sheet_title,doc_type, child, str(unique_id), parent_count, unit))
                cur.execute("""
                    INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (heading_id, sub_heading) 
                    WHERE sub_heading IS NOT NULL 
                    DO UPDATE SET
                        field_data = EXCLUDED.field_data,
                        ordering_index = EXCLUDED.ordering_index;
                """, (str(unique_id), child, row[val_index], parent_count))
                parent_count+=1
                child_count=0
            print("child:" ,child)


def correct_balance_sheet(cashflowTitle,st):
    # try:
    sheet_title="Balance Sheet"
    print(cashflowTitle)
    sheet = workbook[value_to_sheet_title[cashflowTitle]]
    parent = ""
    parent_count = 0
    child_count = 0
    unique_id = None
    unique_id_independent = uuid.uuid4()
    flag=0
    content={2:[3,4,5],3:{6,7,8}}
    parent_index=0
    child_index=0
    # data={}
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue

        if row[0] is None or (isinstance(row[0], str) and row[0].strip().lower() == 'none'):
            continue
        # print("row [0]",row[0])
        if "Abstract".lower() in row[0]:
            continue
        if "Commitments and contingencies".lower() in row[0]:
                continue
            
        # if "equity" in row[0].lower():
        #     continue
        if st==2 and "total" in row[0].lower():
            flag=1
        
        normalized_row = [cell.replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]
        content.append(normalized_row)
        if all(cell is None or cell == '' for cell in normalized_row[1:]):
            
            flag=0
                
            parent = row[0]
            
            #LENGTH EXCEEDING CASE
            if len(parent)>255:
                break
            
            parent_index=index
            print("parent:", parent)
            
            
            
        else:
            child = row[0]
            
            if "Total stockholders' equity".lower() in child.lower() or "Total Shareholders' Equity".lower() in child.lower():
                child_index=index
                break
            if "total" in child and "equity" in child:
                cur.execute("""
                    INSERT INTO metric_headings (ticker, quarter, year, sheet_type, heading, heading_id, ordering_index)
                    VALUES (%s, %s, %s, %s, %s,%s, %s);
                """, (ticker, quarter, year, sheet_title,child, str(unique_id), parent_index))
                cur.execute("""
                    INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (heading_id, sub_heading) 
                    WHERE sub_heading IS NOT NULL 
                    DO UPDATE SET
                        field_data = EXCLUDED.field_data,
                        ordering_index = EXCLUDED.ordering_index;
                """, (str(unique_id), child, row[val_index], parent_count))
                
            if parent != "":
                if flag==1 and "long term" not in child.lower() and "total" not in child.lower():
                        child="Long Term "+child
                
                
                

            else:
                print()
                
                
            print("child:" ,child)
    
    
    
    
    
    
    
    
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue

        if row[0] is None or (isinstance(row[0], str) and row[0].strip().lower() == 'none'):
            continue
        # print("row [0]",row[0])
        
        if "Abstract".lower() in row[0]:
            continue
        if "Commitments and contingencies".lower() in row[0]:
            continue
        if "equity" in row[0].lower() and index!=parent_index and index!=child_index:
            continue
        if st==2 and "total" in row[0].lower():
            flag=1
        
        normalized_row = [cell.replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]
        content.append(normalized_row)
        if all(cell is None or cell == '' for cell in normalized_row[1:]):
            
            flag=0
                
            parent = row[0]
            
            #LENGTH EXCEEDING CASE
            if len(parent)>255:
                break
            
            unique_id = uuid.uuid4()
            print("parent:", parent)
            
            cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s);
            """, (ticker, quarter, year, sheet_title,doc_type, parent, str(unique_id), parent_count))
            
            parent_count += 1
            child_count = 0
        else:
            child = row[0]
            if parent != "":
                if flag==1 and "long term" not in child.lower() and "total" not in child.lower():
                        child="Long Term "+child
                
                cur.execute("""
                    INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (heading_id, sub_heading) 
                    WHERE sub_heading IS NOT NULL 
                    DO UPDATE SET
                        field_data = EXCLUDED.field_data,
                        ordering_index = EXCLUDED.ordering_index;
                """, (str(unique_id), child, row[val_index], child_count))
                
                    
                child_count += 1

            else:
                
                unique_id = uuid.uuid4()
                
                cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s);
            """, (ticker, quarter, year, sheet_title,doc_type, parent, str(unique_id), parent_count))
                cur.execute("""
                    INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (heading_id, sub_heading) 
                    WHERE sub_heading IS NOT NULL 
                    DO UPDATE SET
                        field_data = EXCLUDED.field_data,
                        ordering_index = EXCLUDED.ordering_index;
                """, (str(unique_id), child, row[val_index], parent_count))
                parent_count+=1
                child_count=0
            print("child:" ,child)
            


def define_parent(title):
    sheet = workbook[value_to_sheet_title[title]]
    is_parent = {}
    parent_index = []
    parent = []
    
    for index, row in reversed(list(enumerate(sheet.iter_rows(values_only=True)))):
        normalized_row = [cell.replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]
        if row[0] is not None and  "Abstract".lower() in row[0]:
            parent_index.append(0)
            continue
        if row[0] is not None and  "Commitments and contingencies".lower() in row[0].lower():
            parent_index.append(0)
            continue
        if row[0] is not None and  "Line Items".lower() in row[0].lower():
            parent_index.append(0)
                # print("FOUND LINE ITEMS")
            continue
        if all(cell is None or cell == '' for cell in normalized_row[1:]):
            # Get the last parent_index value or set it to 0 if the list is empty
            last_index = parent_index[-1] if parent_index else 0
            parent_index.append(last_index + 1)
            
            while len(parent)!= 0:
                print("?")
                if parent[-1][1] < parent_index[-1]:
                    x = parent.pop()
                    is_parent[x[0]+1] = index + 1
                else:
                    break
            parent.append([index, parent_index[-1]])
        else:
            parent_index.append(0)
            continue
    # for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    #     print(index,": ",row)
    parent_index.append(-1)
    parent_index.reverse()
    return [parent_index, is_parent]   




def correct_segments(cashflowTitle,sheet_name, unit, revenueFound, temp):
    # try:
        print(cashflowTitle)
        sheet = workbook[value_to_sheet_title[cashflowTitle]]
        parent = ""
        parent_count = 0
        child_count = 0
        unique_id = None
        
        val_index=1
        
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            if index == 1:
                print(row)
                if row[1]==None:
                    val_index=2
                continue
            if row[0] is None or (isinstance(row[0], str) and row[0].strip().lower() == 'none'):
                continue
            # print("row [0]",row[0])
            if len(row[0])>255:
                continue

            if "Line Items".lower() in row[0].lower():
                print("FOUND LINE ITEMS")
                continue

            normalized_row = [cell.replace('\xa0', '').strip() if isinstance(cell, str) else cell for cell in row]

            if all(cell is None or cell == '' for cell in normalized_row[1:]):
                parent = row[0]
                unique_id = uuid.uuid4()
                
                
                
                cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s, %s);
            """, (ticker, quarter, year, sheet_name,doc_type, parent, str(unique_id), parent_count, unit))
                parent_count += 1
                child_count = 0
            else:
                child = row[0]
                print (parent + "-" + child)
                if parent != "":
                    if (parent + "-" + child) in temp:
                        continue
                    cur.execute("""
                        INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (heading_id, sub_heading) 
                        WHERE sub_heading IS NOT NULL 
                        DO UPDATE SET
                            field_data = EXCLUDED.field_data,
                            ordering_index = EXCLUDED.ordering_index;
                    """, (str(unique_id), child, row[val_index], child_count))
                    child_count += 1

                else:
                    print (child + "-" + child) 
                    if (child + "-" + child) in temp:
                        continue
                    if child.lower()=='revenue':
                        if revenueFound[0]==0:
                            revenueFound[0]=1
                        else:
                            continue
                    child_id = uuid.uuid4()
                    cur.execute("""
                INSERT INTO metric_headings (ticker, quarter, year, sheet_type,doc_type , heading, heading_id, ordering_index, unit)
                VALUES (%s, %s, %s, %s, %s, %s,%s, %s, %s);
            """, (ticker, quarter, year, sheet_name,doc_type, child, str(child_id), parent_count, unit))
                    cur.execute("""
                        INSERT INTO metric_data (heading_id, sub_heading, field_data, ordering_index)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (heading_id, sub_heading) 
                        WHERE sub_heading IS NOT NULL 
                        DO UPDATE SET
                            field_data = EXCLUDED.field_data,
                            ordering_index = EXCLUDED.ordering_index;
                    """, (str(child_id), child, row[val_index], parent_count))
                    parent_count+=1
                    child_count=0
    # except Exception as e:
    #     print(e)
        
def match(cashflowMatch,st):
    cashflowTitle = ""
    flag = 0
    for value in sorted_titles:
        # print(value,'->',title)
        for s in cashflowMatch:
            if "parenthetical" in value.lower() :
                continue
            if s.lower() in value.lower():
                flag = 1
                cashflowTitle = value
                # correct_values(cashflowTitle,st)
                return cashflowTitle

def matchSegments(segMatch):
    a=[]
    cashflowTitle = ""
    for value in sorted_titles:
        cashflowTitle = value
        for pattern in segMatch:
            if re.search(pattern, value, re.IGNORECASE):
                # print("Match found in text1!")
                a.append(value)
                break
    return a

def matchIncomestatement(cashflowMatch):
    a = []
    for value in sorted_titles:
        if 'parenthetical' in value.lower():
            continue
        for s in cashflowMatch:
            if s.lower() in value.lower(): 
                a.append(value)
                break
            # print(s)
    return a
def checkUnit (title):
    if "million" in title.lower():
        return "Millions"
    if "billion" in title.lower():
        return "Billions"
    if "thousand" in title.lower():
        return "Thousands"
    return ""
        
def match(cashflowMatch,st):
    cashflowTitle = ""
    flag = 0
    for value in sorted_titles:
        # print(value,'->',title)
        for s in cashflowMatch:
            if "parenthetical" in value.lower() :
                continue
            if s.lower() in value.lower():
                flag = 1
                cashflowTitle = value
                # correct_values(cashflowTitle,st)
                return cashflowTitle

def matchSegments(segMatch):
    a=[]
    cashflowTitle = ""
    for value in sorted_titles:
        cashflowTitle = value
        for pattern in segMatch:
            if re.search(pattern, value, re.IGNORECASE):
                # print("Match found in text1!")
                a.append(value)
                break
    return a

def matchIncomestatement(cashflowMatch):
    a = []
    for value in sorted_titles:
        if 'parenthetical' in value.lower():
            continue
        for s in cashflowMatch:
            if s.lower() in value.lower(): 
                a.append(value)
                break
            # print(s)
    return a


df=pd.read_csv(r"C:\Users\dell\OneDrive\Documents\10-k 10-q extraction\10-k.csv")

tickers={
    # "VRNT",
    # "MSFT","NVDA",
    #      "DHR", "AXL", "TEX", "THR", "AMD", "CMG", "TWLO", "AAPL", "DASH", "HTH",
         "THR",
         #"TEX","AXL","DHR","XOM","MSFT","SNAP","HTH","PEP","CMG","TSLA","USAP","VRNT","INTU","CSCO","AAPL","NVDA","DASH","PG","SYM","TWLO","WOLF"
    # "CSCO"
}
print(len(tickers))

for ticker in tickers:
    
    quarter=""
    year=""
    doc_url=""
    c=0
    for index,row in df.iterrows(): 
        if row[5]==ticker: # for 10-q ticker -> 8 7 6 16                  10k -> 5 2 3 4
            quarter=row[2]
            year=row[3]
            doc_url=row[4]
            print(ticker,quarter,year)
            download_excel(ticker,quarter,year,doc_url)
            print( ticker,quarter,year)
            workbook = load_workbook(f"C:\\Users\\dell\\OneDrive\\Documents\\10-k 10-q extraction\\Financial_Report_{doc_type}{ticker}{quarter}_{year}.xlsx")
            value_to_sheet_title = {}
            sorted_titles = []
            for sheet in workbook.worksheets:
                # print(sheet)
                cell_value = sheet['A1'].value
                if cell_value:
                    cell_value = str(cell_value).strip()
                    if cell_value in value_to_sheet_title:
                        print(f"Duplicate value found: '{cell_value}' for sheets '{value_to_sheet_title[cell_value]}' and '{sheet.title}'")
                    value_to_sheet_title[cell_value] = sheet.title
                    sorted_titles.append(cell_value)
            sorted_titles = sorted(sorted_titles,key=len)
            balanceMatch = ["CONSOLIDATED BALANCE SHEETS","Condensed Balance Sheets", "BALANCE SHEETS - USD", "CONSOLIDATED BALANCE SHEET"]
            incomeMatch = ["consolidated statements of income","statements of income","statement of income","income statements", "income statement", "statements of operations", "statements of operation", "statements of earnings","Earnings statement","STATEMENT OF COMPREHENSIVE INCOME"]
            cashMatch = ["CASH FLOWS STATEMENTS", "Statements of Cash Flows", "CASH FLOWS STATEMENT", "Statement of Cash Flows", "CASH FLOW STATEMENTS", "Statements of Cash Flow"]
            segMatch=[
            r"^(?!.*\bassets?\b)(?!.*\bnarrative\b)(?!.*\btable\b)(?!.*\btables\b)(?!.*\bParenthetical\b)(?=.*\busd\b).*\bsegments?\b.*\bdetails?\b",
            r"^(?!.*\bassets?\b)(?!.*\btable\b)(?!.*\btables\b)(?!.*\bnarrative\b)(?!.*\bParenthetical\b)(?=.*\busd\b).*\bdetails?\b.*\bsegments?\b",
            r"^(?!.*\btable\b)(?!.*\btables\b)(?!.*\bParenthetical\b)(?=.*\busd\b)(?=.*\bDisaggregation\b)(?=.*\bRevenue\b)",
            r"^(?!.*\btable\b)(?!.*\btables\b)(?!.*\bParenthetical\b)(?=.*\bRevenue\b)(?=.*\bRecognition\b)(?=.*\busd\b)",
            r"^(?!.*\btable\b)(?!.*\btables\b)(?!.*\bParenthetical\b)(?=.*\busd\b).*Revenue Classified",
            r"^(?!.*\bassets?\b)(?!.*\bnarrative\b)(?!.*\btable\b)(?!.*\btables\b)(?!.*\bParenthetical\b)(?=.*\busd\b)(?=.*\b(maturit(?:y|ies)|inventor(?:y|ies))\b.*\bdetail(?:s)?\b)",]
            cf=match(cashMatch,1)
            print(cf)
            # x = define_parent(cf)
            
            # correct_values(cf,1,x[0],x[1],checkUnit(cf))
            # ins=matchIncomestatement(incomeMatch)
            # for sheet in ins:
            #     x = define_parent(sheet)
                
            #     correct_values(sheet,0,x[0],x[1],checkUnit(sheet))
            # bs=match(balanceMatch,2)
            # x = define_parent(bs)
            # correct_values(bs,2,x[0],x[1],checkUnit(bs))
            # define_parent(bs)
            segs=matchSegments(segMatch)
            revenueFound=[0]
            temp = {}
            for sheet in segs:
                print(sheet)
                sheet_name="Segments - Revenue"
                if "geograph" in sheet.lower() or "region" in sheet.lower() or "area" in sheet.lower():
                    sheet_name="Segments - Geographical"
                x = define_parent(sheet)
                correct_segments(sheet,sheet_name,checkUnit(sheet),revenueFound, temp)
            print(segs,sep="\n")
            # break

    
    


conn.commit()

cur.close()
conn.close()